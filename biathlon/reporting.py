"""Excel management report for completed training work."""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import COMPONENT_LABELS, COMPONENTS
from .service import analyze_athlete


def _period_frame(frame: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    if frame.empty or "date" not in frame:
        return frame.iloc[0:0].copy()
    dates = pd.to_datetime(frame["date"]).dt.normalize()
    return frame.loc[(dates >= start) & (dates <= end)].copy()


def _style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in writer.book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            width = min(42, max(10, max(len(value) for value in values) + 2))
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def work_report_xlsx_bytes(
    bundle: dict[str, Any],
    start_date: date,
    end_date: date,
) -> bytes:
    """Build a multi-sheet report for the selected completed-work period."""

    start = pd.Timestamp(start_date).normalize()
    end = pd.Timestamp(end_date).normalize()
    if end < start:
        raise ValueError("Крайната дата не може да бъде преди началната дата.")

    today = pd.Timestamp(date.today()).normalize()
    analysis_date = min(end + pd.Timedelta(days=1), today)
    athlete_rows: list[dict[str, Any]] = []
    zone_rows: list[dict[str, Any]] = []
    stress_rows: list[dict[str, Any]] = []
    readiness_rows: list[dict[str, Any]] = []
    activity_frames: list[pd.DataFrame] = []
    daily_frames: list[pd.DataFrame] = []

    athletes = bundle["athletes"].copy()
    for _, athlete in athletes.iterrows():
        athlete_id = str(athlete["athlete_id"])
        athlete_name = str(athlete["name"])
        analysis = analyze_athlete(bundle, athlete_id, as_of=analysis_date, generate_plan=False)
        period_activities = _period_frame(analysis["activity_summaries"], start, end)

        real_columns = [f"real_{component}" for component in COMPONENTS]
        total_real_min = sum(
            float(pd.to_numeric(period_activities.get(column, 0.0), errors="coerce").fillna(0.0).sum())
            if column in period_activities
            else 0.0
            for column in real_columns
        )
        athlete_rows.append(
            {
                "Спортист": athlete_name,
                "Профил": str(athlete.get("profile_name", "")),
                "Тренировъчни дни": int(pd.to_datetime(period_activities["date"]).nunique()) if not period_activities.empty else 0,
                "Тренировки": int(len(period_activities)),
                "Общо часове": round(total_real_min / 60.0, 2),
                "Средна интегрирана готовност": round(float(analysis["integrated"]["integrated_readiness"].mean()), 1),
                "Статус след периода": str(analysis["status"]),
                "Твърд флаг": "Да" if bool(analysis["hard_flag"]) else "Не",
            }
        )

        daily_loads = analysis["daily_loads"].reset_index()
        daily_period = _period_frame(daily_loads, start, end)
        if not daily_period.empty:
            daily_period.insert(0, "Спортист", athlete_name)
            daily_frames.append(daily_period)

        for component in COMPONENTS:
            real_col = f"real_{component}"
            q_col = f"q_{component}"
            e_col = f"e_{component}"
            zone_rows.append(
                {
                    "Спортист": athlete_name,
                    "Компонент": COMPONENT_LABELS.get(component, component),
                    "Реални минути": round(float(pd.to_numeric(period_activities[real_col], errors="coerce").fillna(0).sum()), 1) if real_col in period_activities else 0.0,
                    "Еквивалентни минути Q": round(float(pd.to_numeric(period_activities[q_col], errors="coerce").fillna(0).sum()), 1) if q_col in period_activities else 0.0,
                    "Ефективен товар E": round(float(pd.to_numeric(daily_period[e_col], errors="coerce").fillna(0).sum()), 1) if e_col in daily_period else 0.0,
                }
            )
            stress_rows.append(
                {
                    "Спортист": athlete_name,
                    "Компонент": COMPONENT_LABELS.get(component, component),
                    "7/40 индекс": round(float(analysis["load_stats"].loc[component, "index_7_40"]), 3),
                    "Tref": round(float(analysis["load_stats"].loc[component, "Tref"]), 1),
                    "E7 средно/ден": round(float(analysis["load_stats"].loc[component, "E7_daily"]), 2),
                    "E40 средно/ден": round(float(analysis["load_stats"].loc[component, "E40_daily"]), 2),
                }
            )
            readiness_rows.append(
                {
                    "Спортист": athlete_name,
                    "Компонент": COMPONENT_LABELS.get(component, component),
                    "Товарна readiness %": round(float(analysis["load_readiness"].loc[component, "readiness"]), 1),
                    "Остатъчна умора": round(float(analysis["load_readiness"].loc[component, "fatigue"]), 1),
                    "Дни до практическо възстановяване": round(float(analysis["load_readiness"].loc[component, "days_to_full"]), 1),
                    "Интегрирана готовност %": round(float(analysis["integrated"].loc[component, "integrated_readiness"]), 1),
                    "Твърд флаг": "Да" if bool(analysis["integrated"].loc[component, "hard_flag"]) else "Не",
                    "Основание": str(analysis["integrated"].loc[component, "reason"]),
                }
            )

        if not period_activities.empty:
            export_activities = period_activities.copy()
            export_activities.insert(0, "Спортист", athlete_name)
            activity_frames.append(export_activities)

    athlete_summary = pd.DataFrame(athlete_rows)
    overview = pd.DataFrame(
        [
            {"Показател": "Отчетен период", "Стойност": f"{start.strftime('%d.%m.%Y')} – {end.strftime('%d.%m.%Y')}"},
            {"Показател": "Дата на генериране", "Стойност": date.today().strftime("%d.%m.%Y")},
            {"Показател": "Брой спортисти", "Стойност": len(athlete_summary)},
            {"Показател": "Общо извършени часове", "Стойност": round(float(athlete_summary["Общо часове"].sum()), 2)},
            {"Показател": "Общо тренировки", "Стойност": int(athlete_summary["Тренировки"].sum())},
            {"Показател": "Средна интегрирана готовност", "Стойност": round(float(athlete_summary["Средна интегрирана готовност"].mean()), 1)},
            {"Показател": "Спортисти с твърд флаг", "Стойност": int((athlete_summary["Твърд флаг"] == "Да").sum())},
        ]
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Обобщение", index=False)
        athlete_summary.to_excel(writer, sheet_name="Спортисти", index=False)
        pd.DataFrame(zone_rows).to_excel(writer, sheet_name="Натоварване по зони", index=False)
        pd.DataFrame(stress_rows).to_excel(writer, sheet_name="Стрес и 7-40", index=False)
        pd.DataFrame(readiness_rows).to_excel(writer, sheet_name="Readiness", index=False)
        (pd.concat(daily_frames, ignore_index=True) if daily_frames else pd.DataFrame()).to_excel(writer, sheet_name="Дневен товар", index=False)
        (pd.concat(activity_frames, ignore_index=True) if activity_frames else pd.DataFrame()).to_excel(writer, sheet_name="Тренировки", index=False)
        _style_workbook(writer)
    return output.getvalue()
